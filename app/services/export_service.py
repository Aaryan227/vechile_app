import io
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.db.models.tanker_report import TankerDailyReport

def export_tanker_reports_to_excel(reports: List[TankerDailyReport]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tanker Daily Report"

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = [
        "Sl no", "Date", "U/L point", "RTKM", "Rate",
        "Freight", "Pump", "HSD Ltr", "HSD Rate", "HSD Amt", "Khuraki"
    ]
    
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for idx, r in enumerate(reports, start=1):
        row_data = [
            idx,
            r.report_date.strftime("%d/%m/%Y") if r.report_date else "",
            r.ul_point or "",
            r.rtkm,
            r.rate,
            r.freight,
            r.pump or "",
            r.hsd_ltr,
            r.hsd_rate,
            r.hsd_amount,
            r.khuraki
        ]
        ws.append(row_data)
        row_num = idx + 1
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [4, 5, 6, 8, 9, 10, 11]:
                cell.alignment = align_right

    # Column widths auto-adjust
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_taxes_to_excel(taxes: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Taxes & Charges"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    headers = [
        "Sl No", "Vehicle Number", "Tax Type", "State", "Authority",
        "Period Start", "Period End", "Amount (₹)", "Payment Date", "Due Date",
        "Valid Until", "Status", "Payment Reference", "Challan Number"
    ]

    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for idx, t in enumerate(taxes, start=1):
        veh_num = t.vehicle.vehicle_number if t.vehicle else f"ID: {t.vehicle_id}"
        row_data = [
            idx,
            veh_num,
            t.tax_type.value if hasattr(t.tax_type, 'value') else str(t.tax_type),
            t.state,
            t.tax_authority or "",
            t.period_start.strftime("%d/%m/%Y") if t.period_start else "",
            t.period_end.strftime("%d/%m/%Y") if t.period_end else "",
            t.amount,
            t.payment_date.strftime("%d/%m/%Y") if t.payment_date else "",
            t.due_date.strftime("%d/%m/%Y") if t.due_date else "",
            t.valid_until.strftime("%d/%m/%Y") if t.valid_until else "",
            t.status.value if hasattr(t.status, 'value') else str(t.status),
            t.payment_reference or "",
            t.challan_number or ""
        ]
        ws.append(row_data)
        row_num = idx + 1
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 6, 7, 9, 10, 11, 12]:
                cell.alignment = align_center
            elif col_num == 8:
                cell.alignment = align_right

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
